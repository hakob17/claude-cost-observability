"""Cost Observability sync server.

Receives usage rows from the Claude Code plugin, stores them in SQLite, and
serves an authenticated analytics dashboard with CSV / JSON / Excel export.

Run:  uvicorn app:app --host 0.0.0.0 --port 8321   (from the server/ directory)
"""

import csv
import io
import json
import os
from datetime import datetime, timezone
from pathlib import Path

from fastapi import Depends, FastAPI, HTTPException, Request, Response
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

import db

app = FastAPI(title="Cost Observability", docs_url=None, redoc_url=None)
db.init_db()
if not os.environ.get("COST_OBS_ADMIN_PASSWORD"):
    print("[cost-obs] WARNING: COST_OBS_ADMIN_PASSWORD is not set — dashboard login is disabled")

STATIC_DIR = Path(__file__).parent / "static"
SESSION_COOKIE = "cost_obs_session"


# ------------------------------------------------------------------ auth deps

def bearer_token(request: Request):
    auth = request.headers.get("authorization", "")
    return auth[7:].strip() if auth.lower().startswith("bearer ") else None


def require_admin(request: Request):
    """Dashboard/analytics/token access: the admin session or an admin API token."""
    user = db.session_user(request.cookies.get(SESSION_COOKIE))
    if user:
        return user
    if db.token_role(bearer_token(request)) == "admin":
        return "api-token"
    raise HTTPException(status_code=401, detail="Not authenticated")


def require_ingest(request: Request):
    """Ingest access: an ingest (or admin) API token from the plugin."""
    role = db.token_role(bearer_token(request))
    if role in ("ingest", "admin"):
        return role
    raise HTTPException(status_code=401, detail="Invalid ingest token")


# ------------------------------------------------------------------ auth API

@app.post("/api/login")
async def login(request: Request, response: Response):
    body = await request.json()
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if not db.verify_admin(username, password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = db.create_session(username)
    response.set_cookie(
        SESSION_COOKIE, token, httponly=True, samesite="lax",
        max_age=7 * 86400, path="/",
    )
    return {"ok": True, "user": username}


@app.post("/api/logout")
def logout(request: Request, response: Response):
    db.drop_session(request.cookies.get(SESSION_COOKIE))
    response.delete_cookie(SESSION_COOKIE, path="/")
    return {"ok": True}


@app.get("/api/me")
def me(user: str = Depends(require_admin)):
    return {"user": user}


# ------------------------------------------------------------------ tokens

@app.get("/api/tokens")
def tokens_list(_: str = Depends(require_admin)):
    return {"tokens": db.list_tokens()}


@app.post("/api/tokens")
async def tokens_create(request: Request, _: str = Depends(require_admin)):
    body = await request.json()
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    token = db.create_api_token(name, role="ingest")
    return {"ok": True, "name": name, "token": token}


@app.delete("/api/tokens/{prefix}")
def tokens_delete(prefix: str, _: str = Depends(require_admin)):
    if db.delete_token(prefix) == 0:
        raise HTTPException(status_code=404, detail="token not found")
    return {"ok": True}


# ------------------------------------------------------------------ ingest

@app.post("/api/ingest")
async def ingest(request: Request, _: str = Depends(require_ingest)):
    body = await request.json()
    rows = body if isinstance(body, list) else body.get("rows") or [body]
    clean = []
    for r in rows:
        if not isinstance(r, dict) or not r.get("timestamp"):
            continue
        clean.append({
            "row_id": r.get("row_id"),
            "ts": r.get("timestamp"),
            "session_id": r.get("session_id"),
            "user_email": r.get("user_email"),
            "host": r.get("host"),
            "project": r.get("project"),
            "model": r.get("model"),
            "input_tokens": int(r.get("input_tokens") or 0),
            "output_tokens": int(r.get("output_tokens") or 0),
            "cache_read_tokens": int(r.get("cache_read_tokens") or 0),
            "cache_write_tokens": int(r.get("cache_write_tokens") or 0),
            "turns": int(r.get("turns") or 0),
            "cost_usd": float(r["cost_usd"]) if r.get("cost_usd") is not None else None,
        })
    inserted = db.insert_rows(clean, datetime.now(timezone.utc).isoformat())
    return {"ok": True, "received": len(clean), "inserted": inserted}


# ------------------------------------------------------------------ analytics

@app.get("/api/stats")
def api_stats(days: int = 30, _: str = Depends(require_admin)):
    return db.stats(days=days)


@app.get("/api/rows")
def api_rows(days: int = 30, user: str = None, project: str = None,
             limit: int = 200, order_by: str = "ts", order: str = "desc",
             _: str = Depends(require_admin)):
    return {"rows": db.query_rows(days=days, user=user, project=project,
                                  limit=limit, order_by=order_by, order=order)}


# ------------------------------------------------------------------ export

EXPORT_COLUMNS = [
    "ts", "user_email", "host", "project", "model", "session_id",
    "input_tokens", "output_tokens", "cache_read_tokens", "cache_write_tokens",
    "turns", "cost_usd",
]


@app.get("/api/export")
def export(format: str = "csv", days: int = 30, user: str = None,
           project: str = None, _: str = Depends(require_admin)):
    rows = db.query_rows(days=days, user=user, project=project)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    name = f"claude-usage-{stamp}"

    if format == "json":
        payload = json.dumps([{c: r.get(c) for c in EXPORT_COLUMNS} for r in rows], indent=2)
        return Response(payload, media_type="application/json", headers={
            "Content-Disposition": f'attachment; filename="{name}.json"'})

    if format == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
        return Response(buf.getvalue(), media_type="text/csv", headers={
            "Content-Disposition": f'attachment; filename="{name}.csv"'})

    if format in ("xlsx", "excel"):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Usage"
        ws.append(EXPORT_COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r.get(c) for c in EXPORT_COLUMNS])
        ws.freeze_panes = "A2"
        for i, col in enumerate(EXPORT_COLUMNS, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
                max(12, len(col) + 2)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'})

    raise HTTPException(status_code=400, detail="format must be csv, json, or xlsx")


# ------------------------------------------------------------------ static UI

@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html")


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
